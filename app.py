import requests
from openpyxl import load_workbook
from datetime import date, timedelta,datetime
import csv

csv_1_header = ['Date',
				 'BCB_Commercial_Exports_Total',
				 'BCB_Commercial_Exports_Advances_on_Contracts',
				 'BCB_Commercial_Exports_Payment_Advance',
				 'BCB_Commercial_Exports_Others',
				 'BCB_Commercial_Imports',
				 'BCB_Commercial_Balance',
				 'BCB_Financial_Purchases',
				 'BCB_Financial_Sales',
				 'BCB_Financial_Balance',
				 'BCB_Balance']

csv_2_header = ["Date","BCB_FX_Position"]

source_data_path = {"BCB_Commercial_Exports_Total":"C",
					"BCB_Commercial_Exports_Advances_on_Contracts":"D",
					"BCB_Commercial_Exports_Payment_Advance":"E",
					"BCB_Commercial_Exports_Others":"F",
					"BCB_Commercial_Imports":"G",
					"BCB_Commercial_Balance":"H",
					"BCB_Financial_Purchases":"I",
					"BCB_Financial_Sales":"J",
					"BCB_Financial_Balance":"K",
					"BCB_Balance":"L"}
abr_month = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
			7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
x_range = 1000

def download_file(url):
	"""Download the file from th url"""
	file_name = url.split('/')[-1]
	res = requests.get(url)
	with open(file_name,'wb') as fp:
		fp.write(res.content)
	return file_name

def transformation_of_time_series_type_1(date,file_name):
	"""Looking through the type 1 xls file and fetch values"""
	end_limit = 0
	year_limit = 0
	month_limit = 0
	day_limit = 0

	workbook = load_workbook(file_name)
	worksheet = workbook.active
	
	month,day,year = date.split('/')
	for i in range(1,x_range):
		if worksheet['A'+str(i)].value == 'Memo:':
			end_limit = i
			break
	if end_limit:
		for iy in range(1,end_limit):
			if worksheet['A'+str(iy)].value == int(year):
				year_limit = iy
				break
		if year_limit:
			for im in range(year_limit,end_limit):
				if worksheet['B'+str(im)].value == abr_month[int(month)]:
					month_limit = im
					break

			if month_limit:
				if not isinstance(worksheet['B'+str(month_limit+1)].value,long):
					day_limit = month_limit
				else:
					for idate in range(month_limit,month_limit+30):
						if worksheet['B'+str(idate)].value == int(day):
							day_limit = idate
							break	
					if not day_limit:
						day_limit = month_limit
				return {key:worksheet[val+str(day_limit)].value for key,val in source_data_path.items()}
			else:
				print "Month limit not found"
		else:
			print "Year liit not found"
	else:
		"Wrong file format"

def transformation_of_time_series_type_2(date,file_name):
	"""Looking through the type 2 xls file and fetch values"""
	year_limit = 0
	month_limit = 0
	
	workbook = load_workbook(file_name)
	worksheet = workbook.active
	
	month,day,year = date.split('/')
	for iy in range(1,x_range):
		if worksheet['A'+str(iy)].value == int(year):
			year_limit = iy
			break
	if year_limit:
		for im in range(year_limit,year_limit+13):
			if worksheet['B'+str(im)].value == abr_month[int(month)]:
				month_limit = im
				break

		if month_limit:
			return worksheet["C"+str(month_limit)].value
		else:
			print "Month limit not found"
	else:
		print "Year liit not found"

def create_csv_file_type_1(data):
	"""Create type1 csv file"""
	with open('output_type_1.csv', 'wb') as myfile:
		wr = csv.writer(myfile)
		wr.writerow(csv_1_header)
		for row in data:
			wr.writerow(row)

def create_csv_file_type_2(data):
	"""Create type 2 csv file"""
	with open('output_type_2.csv', 'wb') as myfile:
		wr = csv.writer(myfile)
		wr.writerow(csv_2_header)
		for row in data:
			wr.writerow(row)

def get_the_dates(date_ranges):
	"""date_ranges should be follow the the date formate mm/dd/yyyy-mm/dd/yyyy or mm/dd/yyyy.
	"""
	if '-' in date_ranges:
		start_date_str,end_date_str = date_ranges.split('-')
	else:
		start_date_str = end_date_str = date_ranges

	start_date = datetime.strptime(start_date_str,"%m/%d/%Y")
	end_date = datetime.strptime(end_date_str,"%m/%d/%Y")
	delta = end_date - start_date
	file_name_1 = download_file("http://www.bcb.gov.br/pec/Indeco/Ingl/ie5-24i.xlsx")
	file_name_2 = download_file("http://www.bcb.gov.br/pec/Indeco/Ingl/ie5-26i.xlsx")
	result_1 = []
	result_2 = []

	for i in range(delta.days + 1):
		date = (start_date + timedelta(days=i)).strftime("%m/%d/%Y")
		data = transformation_of_time_series_type_1(date,file_name_1)
		result_1.append([date]+data.values())
	create_csv_file_type_1(result_1)
	print 'type 1 file csv has been generted'

	for i in range(delta.days + 1):
		date = (start_date + timedelta(days=i)).strftime("%m/%d/%Y")
		data = transformation_of_time_series_type_2(date,file_name_2)
		result_2.append([date,data])

	create_csv_file_type_2(result_2)
	print 'type 2 file csv has been generted'

if __name__ == '__main__':
	# type 1 ideal run
	get_the_dates('12/1/2017')
	# type 2 ideal run
	# get_the_dates('3/26/2018-3/28/2018')